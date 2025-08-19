from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q, Count, Sum
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from gestion.models import Alumno, Pago, Horario, Sesion, Asistencia, Profesor, Gasto


@login_required(login_url='login:login')
def reportes(request):
    """Vista principal de reportes"""
    from gestion.models import Alumno, Pago, Asistencia, Gasto
    
    context = {
        'titulo': 'Reportes y Exportaciones',
        'total_alumnos': Alumno.objects.count(),
        'total_pagos': Pago.objects.count(),
        'total_asistencias': Asistencia.objects.count(),
        'total_gastos': Gasto.objects.count(),
    }
    return render(request, 'reportes/reportes.html', context)


def exportar_alumnos_excel(request):
    """Exportar lista de alumnos a Excel"""
    # Obtener parámetros de filtro
    estado = request.GET.get('estado', '')
    curso = request.GET.get('curso', '')
    compartido = request.GET.get('compartido', '')
    
    # Query base
    alumnos = Alumno.objects.all()
    
    # Aplicar filtros
    if estado:
        alumnos = alumnos.filter(activo=estado == 'activo')
    if curso:
        alumnos = alumnos.filter(curso=curso)
    if compartido:
        alumnos = alumnos.filter(es_compartido=compartido == 'compartido')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Apellido', 'DNI', 'Curso', 'Teléfono', 
        'Fecha Nacimiento', 'Dirección', 'Compartido', 'Activo',
        'Fecha Alta', 'Fecha Baja', 'Padre/Madre', 'Tarifa Predeterminada'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, alumno in enumerate(alumnos, 2):
        ws.cell(row=row, column=1, value=alumno.id)
        ws.cell(row=row, column=2, value=alumno.nombre)
        ws.cell(row=row, column=3, value=alumno.apellido)
        ws.cell(row=row, column=4, value=alumno.dni or '')
        ws.cell(row=row, column=5, value=alumno.get_curso_display() if alumno.curso else '')
        ws.cell(row=row, column=6, value=alumno.telefono or '')
        ws.cell(row=row, column=7, value=alumno.fecha_nacimiento.strftime('%d/%m/%Y') if alumno.fecha_nacimiento else '')
        ws.cell(row=row, column=8, value=alumno.direccion or '')
        ws.cell(row=row, column=9, value='Sí' if alumno.es_compartido else 'No')
        ws.cell(row=row, column=10, value='Sí' if alumno.activo else 'No')
        ws.cell(row=row, column=11, value=alumno.fecha_alta.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=12, value=alumno.fecha_baja.strftime('%d/%m/%Y') if alumno.fecha_baja else '')
        ws.cell(row=row, column=13, value=str(alumno.padre) if alumno.padre else '')
        ws.cell(row=row, column=14, value=str(alumno.tarifa_predeterminada) if alumno.tarifa_predeterminada else '')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="alumnos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def exportar_pagos_excel(request):
    """Exportar lista de pagos a Excel"""
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    profesor = request.GET.get('profesor', '')
    
    # Query base
    pagos = Pago.objects.select_related('alumno', 'profesor', 'tarifa').all()
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha__lte=fecha_fin)
        except ValueError:
            pass
    
    if profesor:
        pagos = pagos.filter(profesor__id=profesor)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Alumno', 'Profesor', 'Fecha', 'Concepto', 'Importe Original',
        'Descuento', 'Importe Final', 'Tarifa', 'Número'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, pago in enumerate(pagos, 2):
        ws.cell(row=row, column=1, value=pago.id)
        ws.cell(row=row, column=2, value=str(pago.alumno))
        ws.cell(row=row, column=3, value=str(pago.profesor) if pago.profesor else '')
        ws.cell(row=row, column=4, value=pago.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=pago.concepto)
        ws.cell(row=row, column=6, value=float(pago.importe_original))
        ws.cell(row=row, column=7, value=float(pago.descuento))
        ws.cell(row=row, column=8, value=float(pago.importe_final))
        ws.cell(row=row, column=9, value=str(pago.tarifa) if pago.tarifa else '')
        ws.cell(row=row, column=10, value=pago.numero)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pagos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def exportar_asistencias_excel(request):
    """Exportar lista de asistencias a Excel"""
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    horario = request.GET.get('horario', '')
    asistio = request.GET.get('asistio', '')
    
    # Query base
    asistencias = Asistencia.objects.select_related('alumno', 'sesion', 'sesion__horario').all().order_by('-sesion__inicio')
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            asistencias = asistencias.filter(sesion__inicio__date__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            asistencias = asistencias.filter(sesion__inicio__date__lte=fecha_fin)
        except ValueError:
            pass
    
    if horario:
        asistencias = asistencias.filter(sesion__horario__id=horario)
    
    if asistio:
        asistencias = asistencias.filter(presente=asistio == 'asistio')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Alumno', 'Horario', 'Fecha Sesión', 'Hora Inicio', 'Hora Fin',
        'Presente'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, asistencia in enumerate(asistencias, 2):
        ws.cell(row=row, column=1, value=asistencia.id)
        ws.cell(row=row, column=2, value=str(asistencia.alumno))
        ws.cell(row=row, column=3, value=str(asistencia.sesion.horario))
        ws.cell(row=row, column=4, value=asistencia.sesion.inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=asistencia.sesion.inicio.strftime('%H:%M'))
        ws.cell(row=row, column=6, value=asistencia.sesion.fin.strftime('%H:%M'))
        ws.cell(row=row, column=7, value='Sí' if asistencia.presente else 'No')
    
    # Ajustar ancho de columnas para la primera pestaña
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear segunda pestaña: Horarios y Alumnos Matriculados
    ws2 = wb.create_sheet(title="Horarios")
    
    # Estilos para la segunda pestaña
    header_font2 = Font(bold=True, color="FFFFFF")
    header_fill2 = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment2 = Alignment(horizontal="center", vertical="center")
    
    # Encabezados para horarios
    headers_horarios = [
        'ID', 'Asignatura', 'Profesor', 'Día', 'Hora Inicio', 'Hora Fin',
        'Capacidad', 'Alumnos Matriculados', 'Ocupación %', 'Estado'
    ]
    
    for col, header in enumerate(headers_horarios, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font2
        cell.fill = header_fill2
        cell.alignment = header_alignment2
    
    # Obtener todos los horarios con estadísticas
    from gestion.models import Horario, MatriculaHorario
    horarios = Horario.objects.annotate(
        alumnos_matriculados=Count('matriculas', filter=Q(matriculas__estado='activa'))
    ).order_by('dia_semana', 'hora_inicio')
    
    # Datos de horarios
    for row, horario in enumerate(horarios, 2):
        # Calcular porcentaje de ocupación
        ocupacion = 0
        if horario.capacidad > 0:
            ocupacion = (horario.alumnos_matriculados / horario.capacidad) * 100
        
        ws2.cell(row=row, column=1, value=horario.id)
        ws2.cell(row=row, column=2, value=horario.asignatura)
        ws2.cell(row=row, column=3, value=str(horario.profesor))
        ws2.cell(row=row, column=4, value=horario.get_dia_semana_display())
        ws2.cell(row=row, column=5, value=horario.hora_inicio.strftime('%H:%M'))
        ws2.cell(row=row, column=6, value=horario.hora_fin.strftime('%H:%M'))
        ws2.cell(row=row, column=7, value=horario.capacidad)
        ws2.cell(row=row, column=8, value=horario.alumnos_matriculados)
        ws2.cell(row=row, column=9, value=f"{ocupacion:.1f}%")
        ws2.cell(row=row, column=10, value='Activo' if horario.activo else 'Inactivo')
    
    # Ajustar ancho de columnas para la segunda pestaña
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="asistencias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


def exportar_gastos_excel(request):
    """Exportar lista de gastos a Excel"""
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    categoria = request.GET.get('categoria', '')
    
    # Query base
    gastos = Gasto.objects.all()
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            gastos = gastos.filter(fecha__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            gastos = gastos.filter(fecha__lte=fecha_fin)
        except ValueError:
            pass
    
    if categoria:
        gastos = gastos.filter(categoria=categoria)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Concepto', 'Importe', 'Categoría', 'Fecha', 'Observaciones',
        'Tiene Factura'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, gasto in enumerate(gastos, 2):
        ws.cell(row=row, column=1, value=gasto.id)
        ws.cell(row=row, column=2, value=gasto.concepto)
        ws.cell(row=row, column=3, value=float(gasto.importe))
        ws.cell(row=row, column=4, value=gasto.categoria)
        ws.cell(row=row, column=5, value=gasto.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value=gasto.observaciones or '')
        ws.cell(row=row, column=7, value='Sí' if gasto.has_factura() else 'No')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gastos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response
